import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file['Sheet1']


products_per_supplier = {}
total_inv_per_supplier = {}
products_under_ten = {}
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value
    inventory_price = product_list.cell(product_row, 5)
    # calc number of products per supllier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier[supplier_name]
        products_per_supplier[supplier_name] = current_num_products + 1

    else:
        products_per_supplier[supplier_name] = 1

    #calc total inv per supplier
    if supplier_name in total_inv_per_supplier:
        total_inv_per_supplier.get(supplier_name)
        total_inv_per_supplier[supplier_name] = inventory * price
    else:
        total_inv_per_supplier[supplier_name] = inventory * price
    #products w/ inv < 10
    if inventory < 10:
        products_under_ten[product_num] = inventory
    #add value for total inv price
    inventory_price.value = inventory * price

print(products_per_supplier)
print(total_inv_per_supplier)
print(products_under_ten)
inv_file.save("inventory_w_total_value.xlsx")



